"""
app.py
Servidor Flask — orquestra todos os módulos do RSL Buscador.
"""

import json
import os
import queue
import sys
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path

from flask import Flask, Response, jsonify, render_template, request, send_from_directory

# ─────────────────────────────────────────────
# CONFIGURAÇÃO
# ─────────────────────────────────────────────

BASE_DIR     = Path(__file__).parent
RESULTADOS   = BASE_DIR / "resultados"
HIST_FILE    = BASE_DIR / "historico.json"
RESULTADOS.mkdir(exist_ok=True)

app = Flask(__name__)

# sessões ativas: sid → {queue, event_capes, thread}
_sessoes: dict[str, dict] = {}
_lock = threading.Lock()


# ─────────────────────────────────────────────
# HELPERS — histórico
# ─────────────────────────────────────────────

def _ler_hist() -> list:
    if not HIST_FILE.exists():
        return []
    try:
        return json.loads(HIST_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def _salvar_hist(hist: list):
    HIST_FILE.write_text(json.dumps(hist, ensure_ascii=False, indent=2), encoding="utf-8")


# ─────────────────────────────────────────────
# HELPERS — captura de stdout para SSE
# ─────────────────────────────────────────────

class _QueueWriter:
    """Redireciona print() para uma fila SSE."""
    def __init__(self, q: queue.Queue):
        self._q = q

    def write(self, text: str):
        if text.strip():
            self._q.put({"type": "log", "data": text.rstrip()})

    def flush(self):
        pass


def _sse_stream(sid: str):
    """Gerador SSE: consome a fila da sessão e envia eventos ao navegador."""
    with _lock:
        sess = _sessoes.get(sid)
    if not sess:
        yield "data: {\"type\":\"error\",\"data\":\"Sessão não encontrada\"}\n\n"
        return

    q: queue.Queue = sess["queue"]
    while True:
        try:
            msg = q.get(timeout=30)
            yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
            if msg.get("type") in ("fim", "error", "done", "done-filtrar",
                                   "done-pdfs", "done-obsidian"):
                break
        except queue.Empty:
            yield ": keep-alive\n\n"


# ─────────────────────────────────────────────
# ROTAS PRINCIPAIS
# ─────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/listar-resultados")
def listar_resultados():
    arquivos = sorted(
        [f.name for f in RESULTADOS.iterdir() if f.suffix == ".xlsx"],
        reverse=True,
    )
    return jsonify(arquivos)


@app.route("/download/<path:nome>")
def download(nome):
    return send_from_directory(RESULTADOS, nome, as_attachment=True)


@app.route("/historico")
def historico():
    return jsonify(_ler_hist())


@app.route("/historico/apagar", methods=["POST"])
def apagar_historico():
    """Remove entradas do histórico. Body: {"ids": ["id1","id2",...]} ou {"todos": true}"""
    body = request.get_json(silent=True) or {}
    hist = _ler_hist()

    if body.get("todos"):
        _salvar_hist([])
        return jsonify({"ok": True, "removidos": len(hist)})

    ids = set(body.get("ids", []))
    if not ids:
        return jsonify({"erro": "Nenhum id informado"}), 400

    nova = [h for h in hist if h.get("id") not in ids]
    removidos = len(hist) - len(nova)
    _salvar_hist(nova)
    return jsonify({"ok": True, "removidos": removidos})


@app.route("/stream/<sid>")
def stream(sid):
    return Response(
        _sse_stream(sid),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ─────────────────────────────────────────────
# ROTA — BUSCA
# ─────────────────────────────────────────────

@app.route("/buscar", methods=["POST"])
def buscar():
    body = request.get_json(silent=True) or {}
    descritores_raw = body.get("descritores", "").strip()
    bases           = body.get("bases", [])
    ano_inicio      = int(body.get("ano_inicio", 2020))
    ano_fim         = int(body.get("ano_fim", 2025))

    if not descritores_raw:
        return jsonify({"erro": "Informe ao menos um descritor."}), 400
    if not bases:
        return jsonify({"erro": "Selecione ao menos uma base."}), 400

    descritores = [d.strip() for d in descritores_raw.splitlines() if d.strip()]
    sid = uuid.uuid4().hex
    q   = queue.Queue()
    ev_capes = threading.Event()

    with _lock:
        _sessoes[sid] = {"queue": q, "event_capes": ev_capes}

    def _run():
        old_stdout = sys.stdout
        old_stdin  = sys.stdin
        sys.stdout = _QueueWriter(q)

        class _WebInput:
            def _esperar(self, prompt=""):
                q.put({"type": "aguardar_confirmacao", "data": prompt})
                ev_capes.wait()
                ev_capes.clear()
                return "\n"
            def __call__(self, prompt=""):
                return self._esperar(prompt)
            def readline(self):
                return self._esperar()
            def read(self, n=-1):
                return self._esperar()
        sys.stdin = _WebInput()

        try:
            import importlib
            import buscador_rsl as br

            # Aplica configurações da requisição
            br.ANO_INICIO = ano_inicio
            br.ANO_FIM    = ano_fim

            todos      = []
            contagens  = {d: {} for d in descritores}

            if "scielo" in bases:
                q.put({"type": "etapa", "data": "[1/3] Buscando na SciELO..."})
                from busca_scielo_selenium import buscar_scielo_todos
                scielo_res, scielo_cnts = buscar_scielo_todos(descritores)
                todos.extend(scielo_res)
                for desc, cnt in scielo_cnts.items():
                    contagens.setdefault(desc, {})["SciELO"] = cnt

            if "bdtd" in bases:
                q.put({"type": "etapa", "data": "[2/3] Buscando na BDTD..."})
                for desc in descritores:
                    print(f"  → {desc[:55]}...")
                    res = br.buscar_bdtd(desc)
                    contagens.setdefault(desc, {})["BDTD"] = len(res)
                    todos.extend(res)
                    print(f"     {len(res)} resultado(s)")
                    time.sleep(1.5)

            if "capes" in bases:
                q.put({"type": "etapa", "data": "[3/3] Buscando no Periódicos CAPES..."})
                from busca_capes_selenium import buscar_capes_todos
                capes_res, capes_cnts = buscar_capes_todos(descritores)
                todos.extend(capes_res)
                for desc, cnt in capes_cnts.items():
                    contagens.setdefault(desc, {})["Periódicos CAPES"] = cnt

            q.put({"type": "etapa", "data": "Deduplicando resultados..."})
            unicos, duplicatas = br.deduplicar(todos)

            ts       = datetime.now().strftime("%Y%m%d_%H%M")
            nome_arq = f"RSL_{ts}.xlsx"
            caminho  = RESULTADOS / nome_arq
            br.gerar_excel(todos, unicos, duplicatas, contagens, str(caminho))

            # Salva histórico
            hist = _ler_hist()
            hist.append({
                "id":         uuid.uuid4().hex[:8],
                "timestamp":  datetime.now().strftime("%d/%m/%Y %H:%M"),
                "descritores": descritores,
                "bases":       bases,
                "ano_inicio":  ano_inicio,
                "ano_fim":     ano_fim,
                "arquivo":     nome_arq,
                "total":       len(unicos),
                "duplicatas":  len(duplicatas),
            })
            _salvar_hist(hist)

            q.put({
                "type":       "done",
                "total":      len(unicos),
                "duplicatas": len(duplicatas),
                "data":       nome_arq,
            })

        except Exception as e:
            import traceback
            q.put({"type": "error", "data": traceback.format_exc()})
        finally:
            sys.stdout = old_stdout
            sys.stdin  = old_stdin
            with _lock:
                _sessoes.pop(sid, None)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"id": sid})


@app.route("/confirmar/<sid>", methods=["POST"])
def confirmar(sid):
    with _lock:
        sess = _sessoes.get(sid)
    if sess:
        sess["event_capes"].set()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────
# ROTA — FILTRAR (Ollama · llama3)
# ─────────────────────────────────────────────

@app.route("/filtrar", methods=["POST"])
def filtrar():
    body    = request.get_json(silent=True) or {}
    arquivo = body.get("arquivo", "").strip()
    if not arquivo:
        return jsonify({"erro": "Selecione uma planilha."}), 400

    caminho = RESULTADOS / arquivo
    if not caminho.exists():
        return jsonify({"erro": f"Arquivo não encontrado: {arquivo}"}), 404

    sid = uuid.uuid4().hex
    q   = queue.Queue()
    with _lock:
        _sessoes[sid] = {"queue": q}

    def _run():
        import re as _re
        import json as _json
        old_stdout = sys.stdout
        sys.stdout = _QueueWriter(q)
        try:
            import requests as req_lib
            import openpyxl

            OLLAMA_URL = "http://localhost:11434/api/generate"
            MODELO     = "llama3"

            PROMPT_CRITERIOS = """\
Você é revisor de RSL especializado em vieses algorítmicos e racismo estrutural.

CRITÉRIOS DE INCLUSÃO:
I1 - Recorte temporal: publicado entre 2020 e 2025
I2 - Idioma: português, inglês ou espanhol
I3 - Aderência temática: trata de vieses algorítmicos, racismo algorítmico, discriminação racial automatizada ou IA aplicada a populações negras
I4 - Setor de análise: aborda saúde, segurança pública, mercado de trabalho ou crédito financeiro
I5 - Disponibilidade: texto completo disponível gratuitamente
I6 - Tipo de publicação: artigo, dissertação, tese ou relatório com metodologia explícita

CRITÉRIOS DE EXCLUSÃO:
E1 - Fora do recorte temporal: publicado antes de 2020 ou após 2025
E2 - Sem relação com raça/etnia: aborda vieses algorítmicos sem dimensão racial
E3 - Fora dos setores definidos: foco em entretenimento, educação, agronegócio etc.
E4 - Texto incompleto: texto completo indisponível
E5 - Duplicidade: aparece em mais de uma base
E6 - Inconsistência: título/palavras-chave sugerem aderência mas conteúdo não trata do tema
E7 - Sem metodologia explícita: editorial, resenha ou resumo de evento sem metodologia

Responda SOMENTE com JSON válido (sem texto extra):
{"status": "I", "criterio": "I3, I4", "justificativa": "uma frase explicando a decisão"}

Onde status é:
- "I" = incluir (satisfaz I1+I2+I3 e pelo menos I4 ou I6)
- "E" = excluir (falha em algum critério de exclusão)
- "P" = pendente (incerto — requer leitura completa)\
"""

            wb  = openpyxl.load_workbook(str(caminho))
            aba = wb["Todos os Resultados"] if "Todos os Resultados" in wb.sheetnames else wb.active
            headers = [c.value for c in aba[1]]

            def col(nome):
                try: return headers.index(nome)
                except ValueError: return -1

            idx_titulo    = col("Título")
            idx_resumo    = col("Resumo")
            idx_status    = col("Status (I/E/P)")
            idx_motivo    = col("Motivo Exclusão")
            idx_base      = col("Base")
            idx_descritor = col("Descritor")
            idx_ano       = col("Ano")
            idx_doi       = col("DOI")
            idx_autores   = col("Autores")

            rows  = list(aba.iter_rows(min_row=2))
            total = sum(1 for r in rows if idx_titulo >= 0 and r[idx_titulo].value)
            inc = exc = pen = 0
            processados = 0

            # Coleta dados durante classificação — evita re-leitura de células modificadas
            artigos_classificados = []

            q.put({"type": "etapa", "data": f"Classificando {total} artigos com {MODELO}..."})

            for row in rows:
                titulo = str(row[idx_titulo].value or "").strip() if idx_titulo >= 0 else ""
                if not titulo:
                    continue

                processados += 1
                resumo  = str(row[idx_resumo   ].value or "").strip() if idx_resumo    >= 0 else ""
                ano     = str(row[idx_ano      ].value or "").strip() if idx_ano       >= 0 else ""
                base    = str(row[idx_base     ].value or "").strip() if idx_base      >= 0 else ""
                desc    = str(row[idx_descritor].value or "").strip() if idx_descritor >= 0 else ""
                doi     = str(row[idx_doi      ].value or "").strip() if idx_doi       >= 0 else ""
                autores = str(row[idx_autores  ].value or "").strip() if idx_autores   >= 0 else ""

                prompt = (
                    f"{PROMPT_CRITERIOS}\n\n"
                    f"Título: {titulo}\n"
                    f"Ano: {ano}\n"
                    f"Resumo: {resumo[:800]}\n\n"
                    'Responda APENAS com JSON, exemplo: {"status":"I","criterio":"I3, I4","justificativa":"motivo"}\n'
                    "JSON:"
                )

                classificacao = "P"
                criterio      = ""
                justificativa = ""

                try:
                    resp  = req_lib.post(
                        OLLAMA_URL,
                        json={"model": MODELO, "prompt": prompt, "stream": False},
                        timeout=90,
                    )
                    texto = resp.json().get("response", "").strip()
                    m = _re.search(r'\{[^{}]+\}', texto, _re.DOTALL)
                    if m:
                        try:
                            dados = _json.loads(m.group())
                            classificacao = str(dados.get("status", "P")).strip().upper()
                            if classificacao not in ("I", "E", "P"):
                                classificacao = "P"
                            criterio      = str(dados.get("criterio", "")).strip()
                            justificativa = str(dados.get("justificativa", "")).strip()
                        except _json.JSONDecodeError:
                            pass
                    # fallback: resposta sem JSON
                    if not criterio and not justificativa:
                        # tenta detectar status pela primeira letra
                        for ch in texto.upper():
                            if ch in "IEP":
                                classificacao = ch
                                break
                        # gera justificativa padrão baseada no status
                        justificativa = {
                            "I": "Incluído pela análise automática — revisar critérios manualmente",
                            "E": "Excluído pela análise automática — revisar critérios manualmente",
                            "P": "Pendente — requer leitura completa para decisão",
                        }.get(classificacao, "Pendente — requer leitura completa")
                except Exception as e:
                    print(f"  [Ollama] Erro no artigo {processados}: {e}")
                    classificacao = "P"
                    justificativa = "Erro na análise automática — revisar manualmente"

                # Atualiza células na aba "Todos os Resultados"
                if idx_status >= 0:
                    row[idx_status].value = classificacao
                if idx_motivo >= 0:
                    motivo_txt = f"[{criterio}] {justificativa}" if criterio else justificativa
                    row[idx_motivo].value = motivo_txt

                if classificacao == "I":   inc += 1
                elif classificacao == "E": exc += 1
                else:                      pen += 1

                # Guarda para preencher Tabela 2 depois
                artigos_classificados.append({
                    "base":    base,    "desc":    desc,    "titulo":   titulo,
                    "autores": autores, "ano":     ano,     "doi":      doi,
                    "status":  classificacao,
                    "criterio":      criterio,
                    "justificativa": justificativa,
                })

                q.put({
                    "type":          "progresso_filtrar",
                    "atual":         processados,
                    "total":         total,
                    "classificacao": classificacao,
                    "criterio":      criterio,
                })

            # ── Preenche Tabela 2 com dados coletados (não re-lê células) ────
            if "Pós Filtragem" in wb.sheetnames and artigos_classificados:
                aba2  = wb["Pós Filtragem"]
                hdrs2 = [c.value for c in aba2[1]]

                def col2(nome):
                    try: return hdrs2.index(nome) + 1
                    except ValueError: return -1

                if aba2.max_row > 2:
                    aba2.delete_rows(3, aba2.max_row - 2)

                for rn, a in enumerate(artigos_classificados, 3):
                    motivo_tab2 = f"[{a['criterio']}] {a['justificativa']}" if a["criterio"] else a["justificativa"]
                    for campo, valor in [
                        ("Base",          a["base"]),
                        ("Descritor",     a["desc"]),
                        ("Título",        a["titulo"]),
                        ("Autores",       a["autores"]),
                        ("Ano",           a["ano"]),
                        ("DOI",           a["doi"]),
                        ("Status (I/E/P)", a["status"]),
                        ("Critério",      a["criterio"]),
                        ("Motivo Exclusão", a["justificativa"]),
                    ]:
                        c = col2(campo)
                        if c > 0:
                            aba2.cell(row=rn, column=c, value=valor)

                print(f"  ✅ Tabela 2 preenchida com {len(artigos_classificados)} artigos classificados")

            ts_f = datetime.now().strftime("%Y%m%d_%H%M")
            # evita duplo sufixo _filtrado se já era um arquivo filtrado
            base_nome = arquivo.split("_filtrado_")[0] if "_filtrado_" in arquivo else arquivo.replace(".xlsx", "")
            nome_out  = f"{base_nome}_filtrado_{ts_f}.xlsx"
            wb.save(str(RESULTADOS / nome_out))
            print(f"\n✅ Planilha filtrada salva: {nome_out}")

            q.put({
                "type":     "done-filtrar",
                "incluir":  inc,
                "excluir":  exc,
                "pendente": pen,
                "arquivo":  nome_out,
            })

        except Exception:
            import traceback
            q.put({"type": "error", "data": traceback.format_exc()})
        finally:
            sys.stdout = old_stdout
            with _lock:
                _sessoes.pop(sid, None)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"id": sid})


# ─────────────────────────────────────────────
# ROTA — BAIXAR PDFs
# ─────────────────────────────────────────────

@app.route("/baixar-pdfs", methods=["POST"])
def baixar_pdfs():
    body    = request.get_json(silent=True) or {}
    arquivo = body.get("arquivo", "").strip()
    if not arquivo:
        return jsonify({"erro": "Selecione uma planilha."}), 400

    caminho = RESULTADOS / arquivo
    if not caminho.exists():
        return jsonify({"erro": f"Arquivo não encontrado: {arquivo}"}), 404

    sid = uuid.uuid4().hex
    q   = queue.Queue()
    with _lock:
        _sessoes[sid] = {"queue": q}

    def _run():
        old_stdout = sys.stdout
        sys.stdout = _QueueWriter(q)
        try:
            from baixar_pdfs import processar_planilha
            stats = processar_planilha(str(caminho), callback=print)
            q.put({
                "type":      "done-pdfs",
                "pdfs":      stats.get("pdfs", 0),
                "meta_only": stats.get("meta_only", 0),
                "sem_doi":   stats.get("sem_doi", 0),
            })
        except Exception:
            import traceback
            q.put({"type": "error", "data": traceback.format_exc()})
        finally:
            sys.stdout = old_stdout
            with _lock:
                _sessoes.pop(sid, None)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"id": sid})


# ─────────────────────────────────────────────
# ROTA — EXPORTAR OBSIDIAN
# ─────────────────────────────────────────────

@app.route("/exportar-obsidian", methods=["POST"])
def exportar_obsidian():
    body    = request.get_json(silent=True) or {}
    arquivo = body.get("arquivo", "").strip()
    if not arquivo:
        return jsonify({"erro": "Selecione uma planilha."}), 400

    caminho = RESULTADOS / arquivo
    if not caminho.exists():
        return jsonify({"erro": f"Arquivo não encontrado: {arquivo}"}), 404

    sid = uuid.uuid4().hex
    q   = queue.Queue()
    with _lock:
        _sessoes[sid] = {"queue": q}

    def _run():
        old_stdout = sys.stdout
        sys.stdout = _QueueWriter(q)
        try:
            from exportar_obsidian import exportar
            stats = exportar(str(caminho), callback=print)
            q.put({
                "type":    "done-obsidian",
                "artigos": stats.get("artigos", 0),
                "notas":   stats.get("notas", 0),
            })
        except Exception:
            import traceback
            q.put({"type": "error", "data": traceback.format_exc()})
        finally:
            sys.stdout = old_stdout
            with _lock:
                _sessoes.pop(sid, None)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"id": sid})


# ─────────────────────────────────────────────
# ROTA — PRISMA 2020
# ─────────────────────────────────────────────

def _extrair_dados_prisma(caminho: Path) -> dict:
    """
    Extrai contagens PRISMA 2020 da planilha RSL.

    Fontes de dados:
    ─ "Todos os Resultados":
        · Linhas com Base ∈ {SciELO, BDTD, CAPES} → total único + status I/E/P
        · Bloco de resumo (após linha "Termos Booleanos", até "TOTAL GERAL")
          → brutos por base vindos diretamente da API, antes de qualquer dedup
    ─ Total bruto = soma dos brutos do bloco (SciELO_bruto + BDTD_bruto + CAPES_bruto)
    ─ Duplicatas  = max(0, total_bruto − total_único)  — sempre consistente
    """
    import re as _re
    import openpyxl

    wb = openpyxl.load_workbook(str(caminho), read_only=True, data_only=True)

    dados: dict = {
        "identificacao": {
            "scielo": 0, "bdtd": 0, "capes": 0,
            "total": 0, "unicos": 0, "duplicatas": 0,
        },
        "triagem":       {"apos_dedup": 0},
        "elegibilidade": {"avaliados": 0, "excluidos": 0},
        "inclusao":      {"incluidos": 0, "pendentes": 0},
        "motivos":       {},
        "filtrado":      False,
        "validacoes":    {},
        "avisos":        [],
    }

    BASES_VALIDAS = {"SciELO", "BDTD", "Periódicos CAPES"}

    # ── 1. Total único por base + status (apenas artigos reais) ───────────────
    unicos = 0

    if "Todos os Resultados" in wb.sheetnames:
        aba  = wb["Todos os Resultados"]
        rows = list(aba.iter_rows(min_row=1, values_only=True))
        if rows:
            headers = list(rows[0])

            def _col(nome):
                try:    return headers.index(nome)
                except ValueError: return -1

            idx_base   = _col("Base")
            idx_status = _col("Status (I/E/P)")
            idx_motivo = _col("Motivo Exclusão")

            for row in rows[1:]:
                base = str(row[idx_base] or "").strip() if idx_base >= 0 else ""
                if base not in BASES_VALIDAS:
                    continue   # ignora bloco de resumo e linhas vazias

                unicos += 1

                if idx_status >= 0:
                    s = str(row[idx_status] or "").strip().upper()
                    if s in ("I", "E", "P"):
                        dados["filtrado"] = True
                    if s == "I":
                        dados["inclusao"]["incluidos"] += 1
                    elif s == "E":
                        dados["elegibilidade"]["excluidos"] += 1
                        if idx_motivo >= 0:
                            motivo = str(row[idx_motivo] or "")
                            for code in _re.findall(r'\b[EI]\d\b', motivo):
                                dados["motivos"][code] = dados["motivos"].get(code, 0) + 1
                    elif s == "P":
                        dados["inclusao"]["pendentes"] += 1

    # ── 2. Brutos por base: lidos do bloco de resumo ──────────────────────────
    # Estrutura do bloco (col 0=descritor, 1=SciELO, 2=BDTD, 3=CAPES, 4=Total)
    scielo_bruto = bdtd_bruto = capes_bruto = 0
    bloco_encontrado = False

    if "Todos os Resultados" in wb.sheetnames:
        aba  = wb["Todos os Resultados"]
        rows = list(aba.iter_rows(min_row=1, values_only=True))
        em_bloco = False
        for row in rows[1:]:
            if row[0] == "Termos Booleanos":
                em_bloco = True
                bloco_encontrado = True
                continue
            if row[0] == "TOTAL GERAL":
                break
            if em_bloco and row[0] and isinstance(row[1], (int, float)):
                scielo_bruto += int(row[1] or 0)
                bdtd_bruto   += int(row[2] or 0)
                capes_bruto  += int(row[3] or 0)

    total_bruto = scielo_bruto + bdtd_bruto + capes_bruto

    # Fallback: se o bloco não existir, estima pelo inverso (único + aba Duplicatas)
    if not bloco_encontrado or total_bruto == 0:
        dup_fallback = 0
        if "Duplicatas" in wb.sheetnames:
            aba_dup = wb["Duplicatas"]
            dup_fallback = sum(
                1 for row in aba_dup.iter_rows(min_row=2, values_only=True)
                if any(v for v in row[:3] if v)
            )
        total_bruto  = unicos + dup_fallback
        scielo_bruto = dados["identificacao"]["scielo"]
        bdtd_bruto   = dados["identificacao"]["bdtd"]
        capes_bruto  = dados["identificacao"]["capes"]
        dados["avisos"].append(
            "Bloco de resumo não encontrado — total bruto estimado como único + duplicatas da aba."
        )

    # ── 3. Duplicatas = total_bruto − total_único (nunca negativo) ────────────
    duplicatas = max(0, total_bruto - unicos)

    dados["identificacao"].update({
        "scielo":     scielo_bruto,
        "bdtd":       bdtd_bruto,
        "capes":      capes_bruto,
        "unicos":     unicos,
        "duplicatas": duplicatas,
        "total":      total_bruto,
    })
    dados["triagem"]["apos_dedup"]      = unicos
    dados["elegibilidade"]["avaliados"] = unicos

    # ── 4. Validações ─────────────────────────────────────────────────────────
    ief = (dados["inclusao"]["incluidos"] +
           dados["elegibilidade"]["excluidos"] +
           dados["inclusao"]["pendentes"])
    dados["validacoes"] = {
        "bruto_positivo":    total_bruto > 0,
        "dup_valida":        0 <= duplicatas <= total_bruto,
        "dedup_positivo":    unicos > 0,
        "total_consistente": (total_bruto - duplicatas) == unicos,
        "ief_consistente":   (ief <= unicos) if dados["filtrado"] else None,
    }

    wb.close()
    return dados


@app.route("/prisma")
def prisma():
    return render_template("prisma.html")


@app.route("/prisma/dados", methods=["POST"])
def prisma_dados():
    body    = request.get_json(silent=True) or {}
    arquivo = body.get("arquivo", "").strip()
    if not arquivo:
        return jsonify({"erro": "Selecione uma planilha."}), 400
    caminho = RESULTADOS / arquivo
    if not caminho.exists():
        return jsonify({"erro": f"Arquivo não encontrado: {arquivo}"}), 404
    try:
        dados = _extrair_dados_prisma(caminho)
        return jsonify(dados)
    except Exception:
        import traceback
        return jsonify({"erro": traceback.format_exc()}), 500


# ─────────────────────────────────────────────
# Silencia o Chrome DevTools (evita 404 no log)
# ─────────────────────────────────────────────

@app.route("/.well-known/appspecific/com.chrome.devtools.json")
def chrome_devtools():
    return "", 204


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 55)
    print("  RSL Buscador — http://localhost:5000")
    print("  IMPORTANTE: use terminal externo ao VS Code")
    print("=" * 55)
    app.run(debug=False, host="0.0.0.0", port=5000, threaded=True)
