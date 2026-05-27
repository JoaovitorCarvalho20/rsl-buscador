"""
Buscador RSL - Vieses Algorítmicos e Racismo Estrutural
Busca automática nas bases SciELO (Selenium) e BDTD (API REST)
Importa CSV exportado do Periódicos CAPES
Gera planilha Excel consolidada para RSL
"""

import requests
import time
import re
import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Importa o módulo Selenium para SciELO
from busca_scielo_selenium import buscar_scielo_todos

# ─────────────────────────────────────────────
# CONFIGURAÇÕES
# ─────────────────────────────────────────────

ANO_INICIO = 2020
ANO_FIM = 2025
DESCRITORES = [
    "racismo algoritmo",
    "racismo algorítmico",
    "viés algorítmico",
    "discriminação racial algoritmo",
    "inteligência artificial população negra",
    "algoritmo saúde raça",
    "policiamento preditivo racismo",
    "algoritmo mercado trabalho discriminação",
    "algorithmic discrimination",
    "algorithmic bias race",
    "artificial intelligence racial discrimination",
    "predictive policing racial bias",
    "machine learning racism",
    "algorithmic racism",
]

HEADERS = {"User-Agent": "RSL-TCC-UFVJM/1.0 (pesquisa academica)"}
SESSION = requests.Session()
SESSION.headers.update(HEADERS)

# ─────────────────────────────────────────────
# BDTD — API REST VuFind
# ─────────────────────────────────────────────

BDTD_API = "https://bdtd.ibict.br/vufind/api/v1/search"

def buscar_bdtd(descritor: str) -> list[dict]:
    resultados = []
    termos = re.split(r'\bAND\b', descritor, flags=re.IGNORECASE)
    termos = [t.strip().strip('"') for t in termos if t.strip()]
    query = " ".join(termos)

    page = 1
    page_size = 20
    max_pages = 5

    while page <= max_pages:
        try:
            params = {
                "lookfor": query,
                "type": "AllFields",
                "limit": page_size,
                "page": page,
                "sort": "year desc",
                "filter[]": [f"publishDate:[{ANO_INICIO} TO {ANO_FIM}]"],
                "field[]": [
                    "title", "authors", "publicationDates", "summary",
                    "formats", "publishers", "id", "urls",
                ],
            }
            resp = SESSION.get(BDTD_API, params=params, timeout=20)
            resp.raise_for_status()
            data = resp.json()

            registros = data.get("records", [])
            if not registros:
                break

            for rec in registros:
                titulo = rec.get("title", "").strip()
                if not titulo:
                    continue

                autores_raw = rec.get("authors", {})
                if isinstance(autores_raw, dict):
                    lista_autores = []
                    for grupo in autores_raw.values():
                        if isinstance(grupo, dict):
                            lista_autores.extend(grupo.keys())
                        elif isinstance(grupo, list):
                            lista_autores.extend(grupo)
                    autores = "; ".join(lista_autores)
                elif isinstance(autores_raw, list):
                    autores = "; ".join(str(a) for a in autores_raw)
                else:
                    autores = str(autores_raw)

                pub_dates = rec.get("publicationDates", [])
                ano = str(pub_dates[0]).strip()[:4] if pub_dates else ""
                try:
                    if ano and not (ANO_INICIO <= int(ano) <= ANO_FIM):
                        continue
                except ValueError:
                    pass

                resumo_raw = rec.get("summary", [])
                resumo = resumo_raw[0] if isinstance(resumo_raw, list) and resumo_raw else str(resumo_raw)

                publishers = rec.get("publishers", [])
                instituicao = publishers[0] if isinstance(publishers, list) and publishers else ""

                tipo_raw = str(rec.get("formats", [""])[0]).lower() if rec.get("formats") else ""
                if "disserta" in tipo_raw or "master" in tipo_raw:
                    tipo = "Dissertação"
                elif "tese" in tipo_raw or "doctor" in tipo_raw:
                    tipo = "Tese"
                else:
                    tipo = "Tese/Dissertação"

                urls_raw = rec.get("urls", [])
                if urls_raw and isinstance(urls_raw[0], dict):
                    link = urls_raw[0].get("url", "")
                else:
                    record_id = rec.get("id", "")
                    link = f"https://bdtd.ibict.br/vufind/Record/{record_id}" if record_id else ""

                resultados.append({
                    "base": "BDTD",
                    "descritor": descritor,
                    "titulo": titulo,
                    "autores": autores,
                    "ano": ano,
                    "revista_repositorio": instituicao,
                    "resumo": resumo,
                    "doi": "",
                    "link": link,
                    "tipo": tipo,
                    "status_inclusao": "",
                    "motivo_exclusao": "",
                })

            total = data.get("resultCount", 0)
            if page * page_size >= total:
                break
            page += 1
            time.sleep(1.5)

        except Exception as e:
            print(f"  [BDTD] Erro em '{descritor}': {e}")
            break

    return resultados


# ─────────────────────────────────────────────
# PERIÓDICOS CAPES (importação de CSV)
# ─────────────────────────────────────────────

def importar_capes(caminho_csv: str, descritor: str = "Exportado manualmente") -> list[dict]:
    if not os.path.exists(caminho_csv):
        print(f"  [CAPES] Arquivo não encontrado: {caminho_csv}")
        return []

    df = None
    for enc in ["utf-8-sig", "latin-1", "cp1252"]:
        try:
            df = pd.read_csv(caminho_csv, encoding=enc, on_bad_lines="skip")
            break
        except Exception:
            continue

    if df is None:
        print(f"  [CAPES] Não foi possível ler {caminho_csv}")
        return []

    mapa = {
        "titulo":  ["Title", "Título", "title", "titulo", "TI"],
        "autores": ["Author", "Authors", "Autores", "author", "AU"],
        "ano":     ["Publication Year", "Year", "Ano", "year", "Data", "PY"],
        "revista": ["Source Title", "Journal", "Fonte", "Periódico", "SO"],
        "resumo":  ["Abstract", "Resumo", "abstract", "AB"],
        "doi":     ["DOI", "doi", "DI"],
        "link":    ["URL", "Link", "url", "UR"],
    }

    def achar_col(possiveis):
        for p in possiveis:
            if p in df.columns:
                return p
        return None

    resultados = []
    for _, row in df.iterrows():
        def get(possiveis):
            col = achar_col(possiveis)
            return str(row[col]).strip() if col and pd.notna(row[col]) else ""

        ano = get(mapa["ano"])[:4]
        try:
            if ano and not (ANO_INICIO <= int(ano) <= ANO_FIM):
                continue
        except ValueError:
            pass

        resultados.append({
            "base": "Periódicos CAPES",
            "descritor": descritor,
            "titulo": get(mapa["titulo"]),
            "autores": get(mapa["autores"]),
            "ano": ano,
            "revista_repositorio": get(mapa["revista"]),
            "resumo": get(mapa["resumo"]),
            "doi": get(mapa["doi"]),
            "link": get(mapa["link"]),
            "tipo": "Artigo",
            "status_inclusao": "",
            "motivo_exclusao": "",
        })

    return resultados


# ─────────────────────────────────────────────
# DEDUPLICAÇÃO
# ─────────────────────────────────────────────

def normalizar(texto: str) -> str:
    return re.sub(r'\W+', '', texto.lower())

def deduplicar(registros: list[dict]) -> tuple[list[dict], list[dict]]:
    vistos = {}
    unicos = []
    duplicatas = []
    for r in registros:
        chave = normalizar(r.get("titulo", ""))[:60]
        if not chave:
            unicos.append(r)
            continue
        if chave in vistos:
            dup = r.copy()
            dup["duplicata_de"] = vistos[chave]
            duplicatas.append(dup)
        else:
            vistos[chave] = r.get("titulo", "")
            unicos.append(r)
    return unicos, duplicatas


# ─────────────────────────────────────────────
# GERAÇÃO DA PLANILHA EXCEL
# ─────────────────────────────────────────────

COR_HEADER = "1F3864"
COR_SCIELO = "E8F4FD"
COR_BDTD   = "EAF3DE"
COR_CAPES  = "FFF3CD"

BORDA = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def estilo_header(cell, bg=COR_HEADER):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDA

def estilo_celula(cell, bg="FFFFFF", wrap=False):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border = BORDA
    cell.font = Font(size=9)

def cor_base(base: str) -> str:
    return {"SciELO": COR_SCIELO, "BDTD": COR_BDTD, "Periódicos CAPES": COR_CAPES}.get(base, "FFFFFF")

COLUNAS = [
    ("Base", 14), ("Descritor", 32), ("Título", 48), ("Autores", 28),
    ("Ano", 7), ("Revista/Repositório", 28), ("Tipo", 14),
    ("Resumo", 60), ("DOI", 22), ("Link", 32),
    ("Status (I/E/P)", 14), ("Motivo Exclusão", 28),
]

def escrever_aba_dados(ws, registros, titulo_aba, cor_header=COR_HEADER):
    ws.title = titulo_aba
    ws.row_dimensions[1].height = 30
    for col_idx, (nome, largura) in enumerate(COLUNAS, 1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        estilo_header(cell, cor_header)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}1"

    for row_idx, r in enumerate(registros, 2):
        bg = cor_base(r.get("base", ""))
        # Se DOI presente, usa como link principal
        link = f"https://doi.org/{r['doi']}" if r.get("doi") else r.get("link", "")
        valores = [
            r.get("base",""), r.get("descritor",""), r.get("titulo",""),
            r.get("autores",""), r.get("ano",""), r.get("revista_repositorio",""),
            r.get("tipo",""), r.get("resumo",""), r.get("doi",""), link,
            r.get("status_inclusao",""), r.get("motivo_exclusao",""),
        ]
        for col_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            estilo_celula(cell, bg, wrap=col_idx in (3,4,6,8,10,12))
            if col_idx == 8:
                ws.row_dimensions[row_idx].height = 80

def criar_aba_tabela1(ws, contagens):
    ws.title = "Tabela 1 - Busca Inicial"
    headers = ["Termos Booleanos", "SciELO", "BDTD", "Periódicos CAPES", "Total"]
    larguras = [42, 12, 12, 18, 10]
    for col_idx, (h, w) in enumerate(zip(headers, larguras), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        estilo_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 28

    for row_idx, desc in enumerate(DESCRITORES, 2):
        cnts = contagens.get(desc, {})
        s = cnts.get("SciELO", 0)
        b = cnts.get("BDTD", 0)
        c = cnts.get("Periódicos CAPES", 0)
        for col_idx, val in enumerate([desc, s, b, c, s+b+c], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            estilo_celula(cell, "F9F9F9" if row_idx % 2 == 0 else "FFFFFF")
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="center")

    tot_row = len(DESCRITORES) + 2
    ws.cell(row=tot_row, column=1, value="TOTAL GERAL").font = Font(bold=True, size=10)
    for col_idx in range(2, 6):
        formula = f"=SUM({get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{tot_row-1})"
        cell = ws.cell(row=tot_row, column=col_idx, value=formula)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDA
    ws.freeze_panes = "A2"

def criar_aba_tabela2(ws):
    ws.title = "Pós Filtragem"
    headers = [
        "Base", "Descritor", "Título", "Autores",
        "Ano", "DOI", "Status (I/E/P)", "Critério", "Motivo Exclusão",
    ]
    larguras = [16, 32, 52, 32, 7, 32, 14, 14, 45]
    for col_idx, (h, w) in enumerate(zip(headers, larguras), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        estilo_header(cell, "2E5497")
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 28

    instrucao = ("INSTRUÇÃO: Preenchida automaticamente pela etapa Filtrar (Ollama/llama3). "
                 "Execute a filtragem na interface web para classificar cada artigo.")
    n = len(headers)
    cell_inst = ws.cell(row=2, column=1, value=instrucao)
    cell_inst.font = Font(italic=True, color="666666", size=9)
    ws.merge_cells(f"A2:{get_column_letter(n)}2")
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A1:{get_column_letter(n)}1"

def criar_aba_quadro1(ws):
    ws.title = "Quadro 1 - Estudos Finais"
    colunas = [
        ("Código", 10), ("Título", 50), ("Autor(es) - Ano", 28),
        ("Objetivo Principal", 45), ("População/Contexto", 30),
        ("Metodologia", 30), ("Setor Temático", 22), ("Base de Origem", 16),
        ("Link/DOI", 32),
    ]
    for col_idx, (h, w) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        estilo_header(cell, "7030A0")
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 28

    instrucao = ("INSTRUÇÃO: Preencha após leitura completa dos estudos selecionados. "
                 "Use código E1, E2... igual ao artigo modelo da sua orientadora.")
    ws.cell(row=2, column=1, value=instrucao).font = Font(italic=True, color="666666", size=9)
    ws.merge_cells(f"A2:{get_column_letter(len(colunas))}2")

    for row_idx in range(3, 18):
        ws.cell(row=row_idx, column=1, value=f"E{row_idx-2}").font = Font(bold=True, size=9)
        for col_idx in range(1, len(colunas)+1):
            cell = ws.cell(row=row_idx, column=col_idx)
            estilo_celula(cell, "F3EEF8" if row_idx % 2 == 0 else "FFFFFF", wrap=True)
            ws.row_dimensions[row_idx].height = 55
    ws.freeze_panes = "A3"

def criar_aba_criterios(ws):
    ws.title = "Critérios I-E"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 65

    for col, val in enumerate(["Código", "Critério", "Descrição"], 1):
        estilo_header(ws.cell(row=1, column=col, value=val), "217346")

    inclusao = [
        ("I1", "Recorte temporal", "Estudos publicados entre 2020 e 2025."),
        ("I2", "Idioma", "Publicações em português, inglês ou espanhol."),
        ("I3", "Aderência temática", "Estudos que tratam de vieses algorítmicos, racismo algorítmico ou discriminação racial automatizada."),
        ("I4", "Setor de análise", "Abordam pelo menos um setor: saúde, segurança pública, mercado de trabalho ou crédito financeiro."),
        ("I5", "Disponibilidade", "Texto completo disponível gratuitamente nas plataformas consultadas."),
        ("I6", "Tipo de publicação", "Artigos científicos, dissertações, teses e relatórios com metodologia explícita."),
    ]
    exclusao = [
        ("E1", "Fora do recorte temporal", "Estudos publicados antes de 2020 ou após 2025."),
        ("E2", "Sem recorte racial", "Estudos sobre vieses algorítmicos sem dimensão racial ou étnica."),
        ("E3", "Fora dos setores", "Foco em setores não contemplados no projeto."),
        ("E4", "Texto incompleto", "Texto completo indisponível nas bases consultadas."),
        ("E5", "Duplicidade", "Estudos que aparecem em mais de uma base."),
        ("E6", "Inconsistência", "Título/palavras-chave sugerem aderência, mas conteúdo não trata do tema."),
        ("E7", "Sem metodologia", "Editoriais, resenhas ou resumos sem metodologia descrita."),
    ]

    row = 2
    ws.cell(row=row, column=1, value="INCLUSÃO").font = Font(bold=True, color="217346", size=10)
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="EAF3DE")
    row += 1
    for cod, crit, desc in inclusao:
        for col, val in enumerate([cod, crit, desc], 1):
            estilo_celula(ws.cell(row=row, column=col, value=val), "F4FBF0", wrap=True)
            ws.row_dimensions[row].height = 35
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="EXCLUSÃO").font = Font(bold=True, color="A32D2D", size=10)
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="FCEBEB")
    row += 1
    for cod, crit, desc in exclusao:
        for col, val in enumerate([cod, crit, desc], 1):
            estilo_celula(ws.cell(row=row, column=col, value=val), "FFF5F5", wrap=True)
            ws.row_dimensions[row].height = 35
        row += 1
    ws.freeze_panes = "A2"

def _adicionar_resumo_busca(ws, contagens: dict, n_artigos: int):
    """Adiciona tabela de contagem por descritor/base ao final de 'Todos os Resultados'."""
    inicio = n_artigos + 4          # 2 linhas em branco + 1 de separação
    headers = ["Termos Booleanos", "SciELO", "BDTD", "Periódicos CAPES", "Total"]

    # título do bloco
    titulo_cell = ws.cell(row=inicio, column=1, value="Resumo — Busca Inicial por Descritor")
    titulo_cell.font = Font(bold=True, size=11, color="FFFFFF")
    titulo_cell.fill = PatternFill("solid", fgColor=COR_HEADER)
    titulo_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"A{inicio}:E{inicio}")
    ws.row_dimensions[inicio].height = 26

    # cabeçalho da tabela
    hrow = inicio + 1
    for col_idx, (h, w) in enumerate(zip(headers, [42, 12, 12, 18, 10]), 1):
        cell = ws.cell(row=hrow, column=col_idx, value=h)
        estilo_header(cell)
    ws.row_dimensions[hrow].height = 22

    # linhas de dados
    for offset, desc in enumerate(DESCRITORES, 1):
        row_idx = hrow + offset
        cnts = contagens.get(desc, {})
        s = cnts.get("SciELO", 0)
        b = cnts.get("BDTD", 0)
        c = cnts.get("Periódicos CAPES", 0)
        for col_idx, val in enumerate([desc, s, b, c, s + b + c], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            estilo_celula(cell, "F9F9F9" if offset % 2 == 0 else "FFFFFF")
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="center")

    # total geral
    tot = hrow + len(DESCRITORES) + 1
    ws.cell(row=tot, column=1, value="TOTAL GERAL").font = Font(bold=True, size=10)
    for col_idx in range(2, 6):
        formula = f"=SUM({get_column_letter(col_idx)}{hrow+1}:{get_column_letter(col_idx)}{tot-1})"
        cell = ws.cell(row=tot, column=col_idx, value=formula)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDA


def gerar_excel(todos, unicos, duplicatas, contagens, caminho):
    wb = Workbook()
    wb.remove(wb.active)

    ws_todos = wb.create_sheet()
    escrever_aba_dados(ws_todos, unicos, "Todos os Resultados")
    _adicionar_resumo_busca(ws_todos, contagens, len(unicos))

    criar_aba_tabela2(wb.create_sheet())

    for base in ["SciELO", "BDTD", "Periódicos CAPES"]:
        filtrados = [r for r in unicos if r.get("base") == base]
        if filtrados:
            escrever_aba_dados(wb.create_sheet(), filtrados, base)

    if duplicatas:
        escrever_aba_dados(wb.create_sheet(), duplicatas, "Duplicatas")

    criar_aba_quadro1(wb.create_sheet())
    criar_aba_criterios(wb.create_sheet())

    wb.save(caminho)
    print(f"\n✅ Planilha salva em: {caminho}")


# ─────────────────────────────────────────────
# EXECUÇÃO PRINCIPAL
# ─────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  RSL Buscador — Vieses Algorítmicos e Racismo Estrutural")
    print(f"  Recorte temporal: {ANO_INICIO}–{ANO_FIM}")
    print("=" * 60)

    todos = []
    contagens = {d: {} for d in DESCRITORES}

    # ── SciELO (Selenium) ──
    print("\n[1/3] Buscando na SciELO...")
    from busca_scielo_selenium import buscar_scielo_todos
    scielo_res, scielo_cnts = buscar_scielo_todos(DESCRITORES)
    todos.extend(scielo_res)
    for desc, cnt in scielo_cnts.items():
        contagens[desc]["SciELO"] = cnt

    # ── BDTD ──
    print("\n[2/3] Buscando na BDTD...")
    for desc in DESCRITORES:
        print(f"  → {desc[:55]}...")
        res = buscar_bdtd(desc)
        contagens[desc]["BDTD"] = len(res)
        todos.extend(res)
        print(f"     {len(res)} resultado(s)")
        time.sleep(2)

    # ── Periódicos CAPES (lê RIS já baixados) ──
    print("\n[3/3] Importando Periódicos CAPES (arquivos RIS)...")
    from busca_capes_selenium import parsear_ris, PASTA_RIS
    import glob as glob_module
    arquivos_ris = glob_module.glob(os.path.join(PASTA_RIS, "*.ris"))

    if not arquivos_ris:
        print("  ⚠️  Nenhum arquivo RIS encontrado em capes_exports/")
        print("  Execute busca_capes_selenium.py primeiro para baixar os arquivos.")
    else:
        for arquivo in arquivos_ris:
            nome = os.path.basename(arquivo).replace(".ris", "").replace("_", " ")
            desc_match = next(
                (d for d in DESCRITORES if d[:15].lower() in nome.lower()),
                nome
            )
            registros = parsear_ris(arquivo, desc_match)
            todos.extend(registros)
            contagens[desc_match]["Periódicos CAPES"] = (
                contagens.get(desc_match, {}).get("Periódicos CAPES", 0) + len(registros)
            )
            print(f"  → {os.path.basename(arquivo)}: {len(registros)} registro(s)")

    # Garante contagens zeradas
    for desc in DESCRITORES:
        for base in ["SciELO", "BDTD", "Periódicos CAPES"]:
            contagens[desc].setdefault(base, 0)

    # ── Deduplicação ──
    print(f"\n[Deduplicação] Total bruto: {len(todos)} registros")
    unicos, duplicatas = deduplicar(todos)
    print(f"  Únicos: {len(unicos)} | Duplicatas removidas: {len(duplicatas)}")

    # ── Salvar Excel ──
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    caminho_excel = f"RSL_vieses_algoritmicos_{timestamp}.xlsx"
    gerar_excel(todos, unicos, duplicatas, contagens, caminho_excel)

    # ── Resumo ──
    print("\n── Resumo por base ──")
    for base in ["SciELO", "BDTD", "Periódicos CAPES"]:
        n = sum(contagens[d].get(base, 0) for d in DESCRITORES)
        print(f"  {base:25s}: {n:4d} resultados")
    print(f"  {'Total único':25s}: {len(unicos):4d} resultados")
    print(f"\n✅ Planilha gerada: {caminho_excel}")
    print(f"Próximo passo: abra a planilha e preencha 'Status (I/E/P)':")
    print(f"  I = Incluir | E = Excluir | P = Pendente")


if __name__ == "__main__":
    main()