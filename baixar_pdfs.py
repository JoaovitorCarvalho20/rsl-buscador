"""
baixar_pdfs.py
Lê a planilha Excel do RSL Buscador, tenta baixar o PDF de cada artigo
via Unpaywall (gratuito, sem autenticação além de e-mail) e salva em
~/Meu cofre/RSL/pdfs/. Quando não há PDF disponível, salva os metadados
em JSON para consulta posterior.

Uso standalone:
    python3 baixar_pdfs.py resultados/RSL_20260513_1609.xlsx
"""

import sys
import os
import re
import json
import time
import requests
from pathlib import Path
import openpyxl

# ── configurações ─────────────────────────────────────────────────────────────

UNPAYWALL_EMAIL = "rsl.tcc@pesquisa.br"   # qualquer e-mail válido
PASTA_PDFS      = Path("~/Meu cofre/RSL/pdfs").expanduser()
DELAY_ENTRE_DOIS = 0.25                   # segundos — limite Unpaywall: 10 req/s
TIMEOUT_DOWNLOAD = 45                     # segundos por PDF


# ── helpers ───────────────────────────────────────────────────────────────────

def _nome_seguro(texto: str) -> str:
    """Converte título em nome de arquivo sem caracteres proibidos."""
    return re.sub(r'[\\/:*?"<>|\n\r]', '_', texto).strip()[:100]


def _consultar_unpaywall(doi: str) -> tuple[str | None, str | None]:
    """Retorna (url_pdf, titulo_oa) ou (None, None) se não houver PDF livre."""
    try:
        r = requests.get(
            f"https://api.unpaywall.org/v2/{doi}",
            params={"email": UNPAYWALL_EMAIL},
            timeout=15,
        )
        if r.status_code == 404:
            return None, None
        r.raise_for_status()
        data = r.json()
        loc  = data.get("best_oa_location") or {}
        url  = loc.get("url_for_pdf") or loc.get("url")
        titulo = data.get("title")
        return url, titulo
    except Exception:
        return None, None


def _baixar_arquivo(url: str, destino: Path) -> bool:
    """Faz download de `url` para `destino`. Retorna True se bem-sucedido."""
    try:
        r = requests.get(
            url,
            timeout=TIMEOUT_DOWNLOAD,
            stream=True,
            headers={"User-Agent": "RSL-TCC/1.0 (pesquisa academica)"},
            allow_redirects=True,
        )
        r.raise_for_status()
        content_type = r.headers.get("Content-Type", "")
        if "pdf" not in content_type and "octet-stream" not in content_type:
            return False
        with open(destino, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                f.write(chunk)
        return destino.stat().st_size > 1024  # arquivo vazio = falha
    except Exception:
        return False


def _salvar_meta(caminho: Path, dados: dict):
    """Salva metadados em Markdown com frontmatter YAML (fallback sem PDF)."""
    doi_link = (f"https://doi.org/{dados['doi']}" if dados.get("doi") else dados.get("link", ""))
    esc = lambda t: str(t or "").replace('"', '\\"')
    conteudo = f"""---
titulo: "{esc(dados['titulo'])}"
autores: "{esc(dados.get('autores', ''))}"
ano: {dados.get('ano') or 'null'}
base: "{esc(dados.get('base', ''))}"
doi: "{esc(dados.get('doi', ''))}"
url_pdf: "{esc(dados.get('url_pdf', ''))}"
tags:
  - RSL
  - sem-pdf
---

# {dados['titulo']}

**Autores:** {dados.get('autores') or '—'}
**Ano:** {dados.get('ano') or '—'} | **Base:** {dados.get('base') or '—'}
**DOI:** {dados.get('doi') or '—'}

## Resumo

{dados.get('resumo') or '_Sem resumo disponível._'}

## Acesso

{"[Acessar via DOI](" + doi_link + ")" if doi_link else "_Link não disponível._"}

> ⚠️ PDF não disponível em acesso aberto via Unpaywall.
"""
    caminho.write_text(conteudo, encoding="utf-8")


# ── função principal ──────────────────────────────────────────────────────────

def processar_planilha(caminho_excel: str, callback=print) -> dict:
    """
    Processa todas as linhas da planilha e tenta baixar cada PDF.

    Retorna dict com estatísticas: total, pdfs, meta_only, sem_doi, erros.
    """
    PASTA_PDFS.mkdir(parents=True, exist_ok=True)

    wb  = openpyxl.load_workbook(caminho_excel, data_only=True)
    aba = next(
        (wb[n] for n in ["Todos os Resultados"] + wb.sheetnames if n in wb.sheetnames),
        wb.active,
    )

    # mapeia nomes de colunas → índice (0-based)
    headers = [c.value for c in aba[1]]
    idx = {h: i for i, h in enumerate(headers) if h}

    def cel(row, nome, fallback=0):
        i = idx.get(nome, fallback)
        return str(row[i] or "").strip() if i < len(row) else ""

    stats = {"total": 0, "pdfs": 0, "meta_only": 0, "sem_doi": 0, "erros": 0}

    for row in aba.iter_rows(min_row=2, values_only=True):
        titulo = cel(row, "Título", 2)
        if not titulo:
            continue

        stats["total"] += 1
        doi    = cel(row, "DOI", 8)
        base   = cel(row, "Base", 0)
        autores= cel(row, "Autores", 3)
        ano    = cel(row, "Ano", 4)
        link   = cel(row, "Link", 9)
        resumo = cel(row, "Resumo", 7)

        nome   = _nome_seguro(titulo)
        meta   = {
            "titulo": titulo, "autores": autores, "ano": ano,
            "base": base, "doi": doi, "link": link,
            "resumo": resumo, "url_pdf": None,
        }

        # ── sem DOI: salva só metadados ───────────────────────────────────
        if not doi:
            stats["sem_doi"] += 1
            _salvar_meta(PASTA_PDFS / f"{nome}.md", meta)
            callback(f"[sem DOI] {titulo[:65]}")
            continue

        # ── consulta Unpaywall ────────────────────────────────────────────
        callback(f"[{stats['total']}] {titulo[:55]}...")
        url_pdf, _ = _consultar_unpaywall(doi)
        meta["url_pdf"] = url_pdf or ""

        if url_pdf:
            dest_pdf = PASTA_PDFS / f"{nome}.pdf"
            ok = _baixar_arquivo(url_pdf, dest_pdf)
            if ok:
                stats["pdfs"] += 1
                callback(f"  ✅ PDF salvo  ({dest_pdf.stat().st_size // 1024} KB)")
            else:
                dest_pdf.unlink(missing_ok=True)
                stats["meta_only"] += 1
                callback(f"  ⚠  Download falhou — só metadados")
        else:
            stats["meta_only"] += 1
            callback(f"  📄 Sem acesso aberto — só metadados")

        _salvar_meta(PASTA_PDFS / f"{nome}.md", meta)
        time.sleep(DELAY_ENTRE_DOIS)

    total_desc = stats["pdfs"] + stats["meta_only"] + stats["sem_doi"]
    callback(
        f"\n── Resumo ──\n"
        f"  Total processado : {stats['total']}\n"
        f"  PDFs baixados    : {stats['pdfs']}\n"
        f"  Só metadados     : {stats['meta_only']}\n"
        f"  Sem DOI          : {stats['sem_doi']}\n"
        f"  Pasta            : {PASTA_PDFS}"
    )
    return stats


# ── execução standalone ───────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 baixar_pdfs.py <caminho_planilha.xlsx>")
        sys.exit(1)
    processar_planilha(sys.argv[1])
