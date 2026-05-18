"""
exportar_obsidian.py
Lê a planilha Excel do RSL Buscador e cria notas Markdown no vault do Obsidian.

Estrutura gerada em ~/Meu cofre/RSL/:
  artigos/          — uma nota por artigo com frontmatter YAML
  descritores/      — uma nota índice por descritor
  MOC - RSL.md      — mapa geral de conhecimento com links [[]]

Cada nota de artigo contém [[links]] para:
  - outros artigos do mesmo descritor (teia de conhecimento)
  - a nota do descritor correspondente

Uso standalone:
    python3 exportar_obsidian.py resultados/RSL_20260513_1609.xlsx
"""

import sys
import re
import os
from pathlib import Path
from datetime import datetime
import openpyxl

# ── configurações ─────────────────────────────────────────────────────────────

VAULT_RSL        = Path("~/Meu cofre/RSL").expanduser()
PASTA_ARTIGOS    = VAULT_RSL / "artigos"
PASTA_DESCRITORES= VAULT_RSL / "descritores"


# ── helpers ───────────────────────────────────────────────────────────────────

def _slug(texto: str) -> str:
    """Nome de arquivo seguro para Obsidian (sem [[, ]], #, /, etc.)."""
    return re.sub(r'[\\/:*?"<>|#\[\]\n\r]', '_', str(texto)).strip()[:90]


def _esc(texto) -> str:
    """Escapa aspas duplas para YAML."""
    return str(texto or "").replace('"', '\\"')


# ── leitura da planilha ───────────────────────────────────────────────────────

def _ler_artigos(caminho_excel: str) -> list[dict]:
    wb  = openpyxl.load_workbook(caminho_excel, data_only=True)
    aba = next(
        (wb[n] for n in ["Todos os Resultados"] + wb.sheetnames if n in wb.sheetnames),
        wb.active,
    )

    headers = [c.value for c in aba[1]]
    idx = {h: i for i, h in enumerate(headers) if h}

    def cel(row, nome, fb=0):
        i = idx.get(nome, fb)
        return str(row[i] or "").strip() if i < len(row) else ""

    artigos = []
    for row in aba.iter_rows(min_row=2, values_only=True):
        titulo = cel(row, "Título", 2)
        if not titulo:
            continue
        artigos.append({
            "titulo":    titulo,
            "slug":      _slug(titulo),
            "autores":   cel(row, "Autores", 3),
            "ano":       cel(row, "Ano", 4),
            "base":      cel(row, "Base", 0),
            "descritor": cel(row, "Descritor", 1),
            "tipo":      cel(row, "Tipo", 6),
            "resumo":    cel(row, "Resumo", 7),
            "doi":       cel(row, "DOI", 8),
            "link":      cel(row, "Link", 9),
            "status":    cel(row, "Status (I/E/P)", 10),
            "motivo":    cel(row, "Motivo Exclusão", 11),
        })
    return artigos


# ── geração de notas ──────────────────────────────────────────────────────────

def _nota_artigo(artigo: dict, relacionados: list[dict]) -> str:
    a = artigo
    doi_link  = f"https://doi.org/{a['doi']}" if a["doi"] else a["link"]
    slug_desc = _slug(a["descritor"])

    links_rel = "\n".join(
        f"- [[{x['slug']}|{x['titulo'][:65]}]] ({x['ano']}) — _{x['base']}_"
        for x in relacionados[:15]
    ) or "_Nenhum outro artigo com este descritor._"

    status_label = {"I": "✅ Incluir", "E": "❌ Excluir", "P": "🔶 Pendente"}.get(
        a["status"].upper(), "⬜ Não classificado"
    )

    return f"""---
titulo: "{_esc(a['titulo'])}"
autores: "{_esc(a['autores'])}"
ano: {a['ano'] or 'null'}
base: "{_esc(a['base'])}"
descritor: "{_esc(a['descritor'])}"
tipo: "{_esc(a['tipo'])}"
doi: "{_esc(a['doi'])}"
status: "{a['status']}"
tags:
  - RSL
  - {a['base'].replace(' ', '_').replace('/', '_') or 'sem_base'}
  - {a['tipo'] or 'Artigo'}
---

# {a['titulo']}

**Autores:** {a['autores'] or '—'}
**Ano:** {a['ano'] or '—'} | **Base:** {a['base']} | **Tipo:** {a['tipo']}
**Descritor:** [[Descritor - {slug_desc}]]

## Resumo

{a['resumo'] or '_Sem resumo disponível._'}

## Acesso

{"[Acessar via DOI](" + doi_link + ")" if doi_link else "_Link não disponível._"}

## Classificação RSL

**Status:** {status_label}
{("**Motivo de exclusão:** " + a['motivo']) if a['motivo'] else ""}

## Artigos relacionados (mesmo descritor)

{links_rel}
"""


def _nota_descritor(descritor: str, artigos: list[dict]) -> str:
    slug_desc = _slug(descritor)

    por_status = {"I": [], "E": [], "P": [], "": []}
    for a in artigos:
        por_status.setdefault(a["status"].upper(), por_status[""]).append(a)

    def bloco(label, lista):
        if not lista:
            return ""
        linhas = "\n".join(
            f"- [[{a['slug']}|{a['titulo'][:65]}]] ({a['ano']}) — {a['base']}"
            for a in lista
        )
        return f"\n### {label} ({len(lista)})\n\n{linhas}\n"

    return f"""---
tipo: descritor
descritor: "{_esc(descritor)}"
total: {len(artigos)}
tags:
  - RSL
  - descritor
---

# Descritor: {descritor}

**Total de artigos:** {len(artigos)}

## Artigos
{bloco('✅ Incluídos', por_status.get('I', []))}
{bloco('🔶 Pendentes', por_status.get('P', []) + por_status.get('', []))}
{bloco('❌ Excluídos', por_status.get('E', []))}

## Links rápidos

[[MOC - RSL]] | [[Critérios de Inclusão e Exclusão]]
"""


def _teia_descritores(descritores_map: dict) -> str:
    """Gera seção de teia mostrando conexões entre descritores via artigos compartilhados."""
    # Agrupa artigos por slug para detectar quais aparecem em múltiplos descritores
    slug_descritores: dict[str, set] = {}
    for desc, arts in descritores_map.items():
        for a in arts:
            slug_descritores.setdefault(a["slug"], set()).add(desc)

    # Conexões: pares de descritores que têm artigos em comum
    conexoes: dict[tuple, int] = {}
    for slug, descs in slug_descritores.items():
        descs_list = sorted(descs)
        for i in range(len(descs_list)):
            for j in range(i + 1, len(descs_list)):
                par = (descs_list[i], descs_list[j])
                conexoes[par] = conexoes.get(par, 0) + 1

    if not conexoes:
        return ""

    linhas = "\n".join(
        f"- [[Descritor - {_slug(d1)}|{d1}]] ↔ [[Descritor - {_slug(d2)}|{d2}]] ({n} artigo{'s' if n > 1 else ''} em comum)"
        for (d1, d2), n in sorted(conexoes.items(), key=lambda x: -x[1])[:20]
    )
    return f"\n## Teia de Conexões entre Descritores\n\n{linhas}\n"


def _nota_moc(artigos: list[dict], descritores_map: dict) -> str:
    total_i = sum(1 for a in artigos if a["status"].upper() == "I")
    total_e = sum(1 for a in artigos if a["status"].upper() == "E")
    total_p = sum(1 for a in artigos if a["status"].upper() not in ("I", "E"))

    linhas_desc = "\n".join(
        f"- [[Descritor - {_slug(d)}]] — {len(arts)} artigos"
        for d, arts in sorted(descritores_map.items(), key=lambda x: -len(x[1]))
    )

    bases = {}
    for a in artigos:
        bases[a["base"]] = bases.get(a["base"], 0) + 1
    linhas_bases = "\n".join(f"- **{b}:** {n}" for b, n in sorted(bases.items(), key=lambda x: -x[1]))

    return f"""---
tipo: MOC
atualizado: "{datetime.now().strftime('%Y-%m-%d %H:%M')}"
total_artigos: {len(artigos)}
tags:
  - RSL
  - MOC
---

# RSL — Mapa de Conhecimento

> Vieses Algorítmicos e Racismo Estrutural · 2020–2025

**Atualizado em:** {datetime.now().strftime('%d/%m/%Y às %H:%M')}

## Estatísticas

| | |
|---|---|
| Total de artigos únicos | **{len(artigos)}** |
| ✅ Para incluir | **{total_i}** |
| 🔶 Pendentes | **{total_p}** |
| ❌ Excluídos | **{total_e}** |
| Descritores utilizados | **{len(descritores_map)}** |

## Por base de dados

{linhas_bases}

## Descritores

{linhas_desc}

## Navegação

- [[Critérios de Inclusão e Exclusão]]
- Pasta `artigos/` — notas individuais
- Pasta `descritores/` — índices por termo
{_teia_descritores(descritores_map)}
"""


# ── função principal ──────────────────────────────────────────────────────────

def exportar(caminho_excel: str, callback=print) -> dict:
    """
    Gera todas as notas Markdown no vault do Obsidian.
    Retorna dict com estatísticas: artigos, descritores.
    """
    PASTA_ARTIGOS.mkdir(parents=True, exist_ok=True)
    PASTA_DESCRITORES.mkdir(parents=True, exist_ok=True)

    callback("Lendo planilha...")
    artigos = _ler_artigos(caminho_excel)
    if not artigos:
        callback("⚠ Nenhum artigo encontrado na planilha.")
        return {"artigos": 0, "descritores": 0}

    # Agrupa por descritor
    descritores_map: dict[str, list] = {}
    for a in artigos:
        descritores_map.setdefault(a["descritor"], []).append(a)

    # ── notas de artigos ──────────────────────────────────────────────────────
    callback(f"Gerando {len(artigos)} notas de artigos...")
    slugs_vistos: dict[str, int] = {}
    for a in artigos:
        # desambigua slugs duplicados
        s = a["slug"]
        if s in slugs_vistos:
            slugs_vistos[s] += 1
            a["slug"] = f"{s}_{slugs_vistos[s]}"
        else:
            slugs_vistos[s] = 0

        relacionados = [
            x for x in descritores_map.get(a["descritor"], [])
            if x["slug"] != a["slug"]
        ]
        conteudo = _nota_artigo(a, relacionados)
        (PASTA_ARTIGOS / f"{a['slug']}.md").write_text(conteudo, encoding="utf-8")
        callback(f"  📝 {a['slug'][:70]}")

    # ── notas de descritores ──────────────────────────────────────────────────
    callback(f"Gerando {len(descritores_map)} notas de descritores...")
    for desc, arts in descritores_map.items():
        conteudo = _nota_descritor(desc, arts)
        (PASTA_DESCRITORES / f"Descritor - {_slug(desc)}.md").write_text(
            conteudo, encoding="utf-8"
        )

    # ── MOC principal ─────────────────────────────────────────────────────────
    callback("Gerando MOC principal...")
    (VAULT_RSL / "MOC - RSL.md").write_text(
        _nota_moc(artigos, descritores_map), encoding="utf-8"
    )

    callback(
        f"\n── Resumo ──\n"
        f"  Artigos exportados : {len(artigos)}\n"
        f"  Descritores        : {len(descritores_map)}\n"
        f"  Vault              : {VAULT_RSL}"
    )
    return {"artigos": len(artigos), "descritores": len(descritores_map)}


# ── execução standalone ───────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 exportar_obsidian.py <caminho_planilha.xlsx>")
        sys.exit(1)
    exportar(sys.argv[1])
